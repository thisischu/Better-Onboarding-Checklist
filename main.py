import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import subprocess
import platform
from openpyxl.styles import Font
from config import input_file, output_file, start_date  # Importing from config.py

# Read and process the Excel file from the sheet named "FTE Hire"
df = pd.read_excel(input_file, sheet_name='FTE Hire')

# Strip any extra spaces from column names
df.columns = df.columns.str.strip()

# Function to generate email alias for SPOC
def generate_spoc_email_alias(full_name):
    names = full_name.split()
    if len(names) < 2:
        return ""  
    first_name = names[0].lower()
    last_name_initial = names[-1][0].lower()  # Get the first letter of the last name
    email_alias = f"{first_name}{last_name_initial}@better.com"
    return email_alias

# Function to generate email alias for Neo
def generate_neo_email_alias(full_name):
    names = full_name.split()
    if len(names) < 2:
        return ""  
    first_initial = names[0][0].lower()  # Get the first letter of the first name
    last_name = names[-1].lower()  # Use full last name
    email_alias = f"{first_initial}_{last_name}@neopoweredbybetter.com"
    return email_alias

# Function to extract username (remove "@better.com")
def extract_username(work_email):
    if pd.isna(work_email) or not isinstance(work_email, str):
        return ""  
    return work_email.split('@')[0]

# Main function to create the output Excel sheet
def create_excel(start_date, input_df, output_file):
    # Filter data based on the specified start date
    start_date = pd.to_datetime(start_date, format="%m/%d/%Y", errors='coerce')
    filtered_data = input_df[pd.to_datetime(input_df['Start Date'], errors='coerce') == start_date]

    # Correct the column names based on the actual file's structure
    columns_needed = ['Candidate', 'Personal Email', 'Better Email', 'Job', 'Department', 'Office', 'Start Date', 'Office Location (for OL)', 'TimeZone']
    selected_data = filtered_data[columns_needed]

    # Create a new DataFrame for the output Excel sheet
    output_data = pd.DataFrame({
        'Full Name': selected_data['Candidate'],
        'First Name': selected_data['Candidate'].apply(lambda name: name.split()[0] if pd.notna(name) else ''),  # Extract first name
        'Personal Email': selected_data['Personal Email'],
        'Username': selected_data['Better Email'].apply(extract_username),  
        'Better Email': selected_data['Better Email'],
        'Temporary Password': '',  
        'Title': selected_data['Job'],
        'Department': selected_data['Department'],
        'Front Setup Needed': selected_data['Department'].apply(lambda dept: 'Yes' if dept == 'SPOC' else 'N/A'),
        #'Email Alias Paragraph': selected_data['Department'].apply(lambda dept: 'This is your External Email Alias, please do not refer to this Alias email address until we meet in the IT Onboarding Session:' if dept == 'SPOC' else ''),
        'Email Alias': selected_data.apply(lambda row: generate_spoc_email_alias(row['Candidate']) if row['Department'] == 'SPOC' else 
                                           generate_neo_email_alias(row['Candidate']) if row['Department'] == 'Neo' else 'N/A', axis=1),  # Generate email alias for SPOC and Neo, or 'N/A'
        'Location': selected_data['Office Location (for OL)'],
        'Start Date': pd.to_datetime(selected_data['Start Date'], errors='coerce').dt.strftime('%B %d, %Y'), 
        'Start Time': selected_data['TimeZone'].apply(lambda tz: '10:30 AM EST' if 'EST' in str(tz) else '12:30 AM PST' if 'PST' in str(tz) else ''),  # Set time based on TimeZone
        'Monitor FedEx Tracking': '',
        'WFH Bundle FedEx Tracking': '', 
        'Laptop FedEx Tracking': '',
        'Tracking Status': ''  
    })

    # Generate a sheet name based on the start date
    if not output_data.empty:
        sheet_name = f'NH_{start_date.strftime("%B_%d_%Y")}'
    else:
        sheet_name = 'Sheet_NoStartDate'

    # Load the existing workbook or create a new one if it doesn't exist
    try:
        workbook = openpyxl.load_workbook(output_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Check if the sheet already exists, and if so, remove it
    if sheet_name in workbook.sheetnames:
        std = workbook[sheet_name]
        workbook.remove(std)

    # Add a new sheet to the workbook
    sheet = workbook.create_sheet(title=sheet_name)

    # Write DataFrame to the new sheet
    for r_idx, row in enumerate(dataframe_to_rows(output_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            # Apply border around each cell
            cell.border = Border(left=Side(border_style='thin'),
                                 right=Side(border_style='thin'),
                                 top=Side(border_style='thin'),
                                 bottom=Side(border_style='thin'))

    # Define the styles
    header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  
    bold_font = Font(size=14, bold=True)  
    regular_font = Font(size=14)  

    # Apply styles to the header row
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.font = bold_font  

    # Adjust column widths, row heights, and center align all cells
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
                cell.alignment = center_alignment  
                cell.font = regular_font  
                cell.border = Border(left=Side(border_style='thin'),
                                     right=Side(border_style='thin'),
                                     top=Side(border_style='thin'),
                                     bottom=Side(border_style='thin'))
        sheet.column_dimensions[column_letter].width = max_length + 5 

    # Set row height and wrap text for each row
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment  
            cell.font = regular_font  
            cell.border = Border(left=Side(border_style='thin'),
                                 right=Side(border_style='thin'),
                                 top=Side(border_style='thin'),
                                 bottom=Side(border_style='thin'))
        sheet.row_dimensions[row[0].row].height = 60  # Set row height to 60 pixels

    # Hide columns that are empty
    for col in sheet.columns:
        if all(cell.value is None for cell in col):
            sheet.column_dimensions[col[0].column_letter].hidden = True

    # Hide rows that are empty
    for row in sheet.iter_rows():
        if all(cell.value is None for cell in row):
            sheet.row_dimensions[row[0].row].hidden = True 

    # Save the workbook with the new sheet and formatting
    workbook.save(output_file)

    print(f"New Excel sheet '{sheet_name}' created and formatted successfully in '{output_file}'!")

    # Open the output file
    open_excel_file(output_file)

# Define the open function based on the operating system
def open_excel_file(file_path):
    if platform.system() == 'Darwin':  # macOS
        subprocess.call(['open', file_path])
    elif platform.system() == 'Windows':  # Windows
        import os 
        os.startfile(file_path)
    else:  # Linux
        subprocess.call(['xdg-open', file_path])

# Execute the main function
create_excel(start_date, df, output_file)
